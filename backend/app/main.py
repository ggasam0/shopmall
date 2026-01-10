from pathlib import Path

from fastapi import Depends, FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from openpyxl import load_workbook
from sqlmodel import Session, select

from app.config import AUTH_CONFIG
from app.db import get_session, init_db
from app.models import AuthAccount, Order, Product, User
from app.schemas import (
    AuthLoginRequest,
    AuthLoginResponse,
    DashboardSummary,
    DistributorSummary,
    OrderRead,
    OrderStatusUpdate,
    PhoneLoginRequest,
    ProductBulkCreate,
    ProductCreate,
    ProductRead,
    UserRead,
)

app = FastAPI(title="Fireworks Mall API")

IMAGE_DIR = Path(__file__).resolve().parent / "images"
IMAGE_DIR.mkdir(exist_ok=True)
app.mount("/images", StaticFiles(directory=str(IMAGE_DIR)), name="images")

PRODUCT_EXCEL_PATH = Path(__file__).resolve().parent / "商品.xlsx"


def load_products_from_excel(session: Session) -> bool:
    if not PRODUCT_EXCEL_PATH.exists():
        return False

    workbook = load_workbook(PRODUCT_EXCEL_PATH, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    header_map = {header: index for index, header in enumerate(headers)}
    required_headers = ["商品名称", "类别", "价格", "图片名称"]
    if not all(header in header_map for header in required_headers):
        return False

    existing_names = set(session.exec(select(Product.name)).all())
    created = False

    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = row[header_map["商品名称"]]
        category = row[header_map["类别"]]
        price = row[header_map["价格"]]
        image_name = row[header_map["图片名称"]]
        if not name or not category or price is None:
            continue
        if name in existing_names:
            continue
        image_url = f"/images/{image_name}" if image_name else ""
        session.add(
            Product(
                name=str(name).strip(),
                category=str(category).strip(),
                price=float(price),
                image_url=image_url,
            )
        )
        created = True

    if created:
        session.commit()
    return created

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def on_startup() -> None:
    init_db()
    with get_session() as session:
        for account in AUTH_CONFIG:
            user_payload = account["user"]
            user = session.exec(select(User).where(User.phone == user_payload["phone"])).first()
            if not user:
                user = User(
                    name=user_payload["name"],
                    phone=user_payload["phone"],
                    role=account["role"],
                    pickup_address=user_payload.get("pickup_address"),
                )
                session.add(user)
                session.commit()
                session.refresh(user)
            else:
                user.pickup_address = user_payload.get("pickup_address")
                session.add(user)
                session.commit()

            existing_account = session.exec(
                select(AuthAccount).where(AuthAccount.username == account["username"])
            ).first()
            if not existing_account:
                session.add(
                    AuthAccount(
                        username=account["username"],
                        password=account["password"],
                        role=account["role"],
                        user_id=user.id,
                    )
                )
                session.commit()

        has_products = bool(session.exec(select(Product)).first())
        loaded_from_excel = load_products_from_excel(session)
        if not has_products and not loaded_from_excel:
            session.add(
                Product(
                    name="夜景礼花套装",
                    category="夜景礼花",
                    price=298.0,
                    image_url="https://images.unsplash.com/photo-1509228468518-180dd4864904",
                    tags="夜景,礼花,套装",
                    is_featured=True,
                )
            )
            session.add(
                Product(
                    name="节庆鞭炮",
                    category="纸炮",
                    price=88.0,
                    image_url="https://images.unsplash.com/photo-1509228468518-180dd4864904",
                    tags="节庆,鞭炮",
                )
            )
            session.commit()

        if not session.exec(select(Order)).first():
            distributor_user = session.exec(
                select(User).where(User.role == "distributor")
            ).first()
            if distributor_user:
                session.add(
                    Order(
                        user_id=distributor_user.id,
                        status="待发货",
                        total=298.0,
                    )
                )
                session.commit()


@app.post("/auth/login", response_model=AuthLoginResponse)
def login(payload: AuthLoginRequest, session: Session = Depends(get_session)) -> AuthLoginResponse:
    account = session.exec(
        select(AuthAccount).where(
            AuthAccount.username == payload.username,
            AuthAccount.password == payload.password,
        )
    ).first()
    if not account or not account.user_id:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    user = session.get(User, account.user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return AuthLoginResponse(role=account.role, user_id=user.id, name=user.name)


@app.post("/auth/phone", response_model=UserRead)
def phone_login(
    payload: PhoneLoginRequest, session: Session = Depends(get_session)
) -> UserRead:
    user = session.exec(select(User).where(User.phone == payload.phone)).first()
    if not user:
        user = User(name="手机用户", phone=payload.phone, role="customer")
        session.add(user)
        session.commit()
        session.refresh(user)
    return user


@app.get("/users", response_model=list[UserRead])
def list_users(session: Session = Depends(get_session)) -> list[UserRead]:
    return session.exec(select(User)).all()


@app.get("/products", response_model=list[ProductRead])
def list_products(session: Session = Depends(get_session)) -> list[ProductRead]:
    return session.exec(select(Product)).all()


@app.post("/products", response_model=ProductRead)
def create_product(
    payload: ProductCreate, session: Session = Depends(get_session)
) -> ProductRead:
    product = Product(**payload.model_dump())
    session.add(product)
    session.commit()
    session.refresh(product)
    return product


@app.post("/products/bulk", response_model=list[ProductRead])
def create_products_bulk(
    payload: ProductBulkCreate, session: Session = Depends(get_session)
) -> list[ProductRead]:
    products = [Product(**item.model_dump()) for item in payload.products]
    session.add_all(products)
    session.commit()
    for product in products:
        session.refresh(product)
    return products


@app.get("/orders", response_model=list[OrderRead])
def list_orders(session: Session = Depends(get_session)) -> list[OrderRead]:
    return session.exec(select(Order)).all()


@app.get("/users/{user_id}/orders", response_model=list[OrderRead])
def list_user_orders(
    user_id: int, session: Session = Depends(get_session)
) -> list[OrderRead]:
    return session.exec(select(Order).where(Order.user_id == user_id)).all()


@app.patch("/orders/{order_id}", response_model=OrderRead)
def update_order_status(
    order_id: int,
    payload: OrderStatusUpdate,
    session: Session = Depends(get_session),
) -> OrderRead:
    order = session.get(Order, order_id)
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    order.status = payload.status
    session.add(order)
    session.commit()
    session.refresh(order)
    return order


@app.get("/admin/summary", response_model=DashboardSummary)
def admin_summary(session: Session = Depends(get_session)) -> DashboardSummary:
    total_sales = sum(order.total for order in session.exec(select(Order)).all())
    pending_orders = len(
        [order for order in session.exec(select(Order)).all() if order.status != "已完成"]
    )
    active_distributors = len(
        [user for user in session.exec(select(User)).all() if user.role == "distributor"]
    )
    featured_products = len(
        [product for product in session.exec(select(Product)).all() if product.is_featured]
    )
    return DashboardSummary(
        total_sales=total_sales,
        pending_orders=pending_orders,
        active_distributors=active_distributors,
        featured_products=featured_products,
    )


@app.get("/distributor/{user_id}/summary", response_model=DistributorSummary)
def distributor_summary(
    user_id: int, session: Session = Depends(get_session)
) -> DistributorSummary:
    user = session.get(User, user_id)
    if not user or user.role != "distributor":
        raise HTTPException(status_code=404, detail="Distributor not found")
    orders = session.exec(select(Order).where(Order.user_id == user_id)).all()
    total_orders = len(orders)
    commission = sum(order.total for order in orders) * 0.15
    return DistributorSummary(
        distributor_id=user.id,
        name=user.name,
        pickup_address=user.pickup_address,
        total_orders=total_orders,
        commission=commission,
        wallet_balance=1200.0,
        coupons=3,
        points=180,
    )
